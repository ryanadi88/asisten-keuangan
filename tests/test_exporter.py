"""
Unit tests for 1-Click Excel multi-sheet exporter.
"""

import os
import io
import pytest
import openpyxl
from datetime import datetime

from database.sqlite_db import SQLiteBackend
from database.models import TransactionCreate, TransactionType, Category, InvoiceCreate, InvoiceStatus
from tools.exporter import ExcelExporter


@pytest.mark.asyncio
async def test_excel_export_generation(tmp_path):
    db_path = str(tmp_path / "test_export.db")
    storage = SQLiteBackend(db_path=db_path)
    await storage.init_db()
    exporter = ExcelExporter(storage=storage)

    user_id = 450

    # Seed Transaction
    tx = TransactionCreate(
        user_id=user_id,
        timestamp=datetime(2026, 8, 10),
        type=TransactionType.INCOME,
        category=Category.BUFFER,
        amount=10_000_000.0,
        source_or_merchant="Client XYZ",
    )
    await storage.add_transaction(tx)

    # Seed Invoice
    inv = InvoiceCreate(
        user_id=user_id,
        client_name="PT Sentosa",
        project_title="Design System",
        amount=8_000_000.0,
        due_date="2026-08-25",
        status=InvoiceStatus.UNPAID,
    )
    await storage.add_invoice(inv)

    excel_bytes = await exporter.export_financial_workbook(user_id=user_id)
    assert excel_bytes is not None
    assert len(excel_bytes) > 1000

    # Load and verify sheets using openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    sheet_names = wb.sheetnames

    assert "Ringkasan Finansial" in sheet_names
    assert "Histori Transaksi" in sheet_names
    assert "Daftar Invoice" in sheet_names

    ws_tx = wb["Histori Transaksi"]
    assert ws_tx.max_row >= 2  # Header + 1 transaction

    ws_inv = wb["Daftar Invoice"]
    assert ws_inv.max_row >= 2  # Header + 1 invoice


def test_pdf_cheatsheet_generation(tmp_path):
    """Test generating PDF Cheatsheet file."""
    from tools.generate_cheatsheet_pdf import generate_cheatsheet_pdf
    pdf_out = str(tmp_path / "test_cheatsheet.pdf")
    res_path = generate_cheatsheet_pdf(pdf_out)
    assert res_path == pdf_out
    assert os.path.exists(pdf_out)
    assert os.path.getsize(pdf_out) > 5000


@pytest.mark.asyncio
async def test_pdf_financial_statement_generation(tmp_path):
    """Test generating PDF financial statement bytes."""
    from tools.pdf_exporter import pdf_financial_exporter
    from database.models import MonthlySummary, UserSettings

    summary = MonthlySummary(
        month_year="2026-08",
        user_id=1,
        total_income=15_000_000.0,
        total_expense=4_500_000.0,
        target_salary=10_000_000.0,
        actual_salary_drawn=10_000_000.0,
        buffer_fund_balance=18_000_000.0,
        emergency_fund=5_000_000.0,
        investment_total=2_000_000.0,
        tax_reserve=1_500_000.0,
    )
    settings = UserSettings(user_id=1)
    pdf_bytes = pdf_financial_exporter.generate_statement_pdf(
        user_id=1,
        month_year="2026-08",
        summary=summary,
        settings=settings,
        transactions=[],
        goals=[],
        subscriptions=[],
    )
    assert pdf_bytes is not None
    assert len(pdf_bytes) > 2000
    assert pdf_bytes.startswith(b"%PDF")

