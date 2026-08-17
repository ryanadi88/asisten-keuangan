"""
1-Click Excel (.xlsx) and CSV Financial Exporter.
Generates multi-sheet formatted financial workbooks in memory.
"""

import io
import logging
from datetime import datetime
from typing import List, Optional

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import format_currency, get_current_month_year
from database.base import StorageBackend
from database import get_storage
from database.models import Transaction, Invoice, MonthlySummary, UserSettings

logger = logging.getLogger(__name__)


class ExcelExporter:
    def __init__(self, storage: Optional[StorageBackend] = None):
        self.storage = storage or get_storage()

    async def export_financial_workbook(self, user_id: int) -> bytes:
        """
        Generate a complete 3-sheet Excel workbook (.xlsx) containing:
        1. Ringkasan Finansial
        2. Histori Transaksi
        3. Daftar Invoice & Piutang
        """
        m_y = get_current_month_year()
        summary = await self.storage.get_monthly_summary(m_y, user_id)
        user_settings = await self.storage.get_user_settings(user_id)
        transactions = await self.storage.get_transactions(user_id=user_id, limit=1000)
        invoices = await self.storage.get_invoices(user_id=user_id, limit=1000)

        # 1. Create openpyxl Workbook
        wb = openpyxl.Workbook()
        # Remove default sheet
        default_sheet = wb.active

        # Style tokens
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        sub_header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        bold_font = Font(name="Arial", size=10, bold=True)
        regular_font = Font(name="Arial", size=10)
        border_thin = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )

        # --- SHEET 1: Ringkasan Finansial ---
        ws_sum = wb.create_sheet(title="Ringkasan Finansial")
        ws_sum.views.sheetView[0].showGridLines = True

        ws_sum.append(["FREELANCE AI FINANCIAL ENGINE — LAPORAN REKAPITULASI"])
        ws_sum.append([f"Bulan / Periode: {m_y} | Export: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
        ws_sum.append([])

        ws_sum.cell(row=1, column=1).font = Font(name="Arial", size=14, bold=True, color="1E3A8A")
        ws_sum.cell(row=2, column=1).font = Font(name="Arial", size=10, italic=True, color="64748B")

        summary_rows = [
            ("Gross Total Income", summary.total_income),
            ("Total Expenses (Pengeluaran)", summary.total_expense),
            ("Net Cashflow (Tabungan Bersih)", summary.net_savings),
            ("Target Gaji Pokok Bulanan", summary.target_salary),
            ("Gaji Ditarik Bulan Ini", summary.actual_salary_drawn),
            ("Cadangan Pajak & Ops (10%)", summary.tax_reserve),
            ("Saldo Buffer Fund (Smoothing Pool)", summary.buffer_fund_balance),
            ("Dana Darurat", summary.emergency_fund),
            ("Portofolio Investasi", summary.investment_total),
            ("Buffer Runway Keamanan (Bulan)", summary.buffer_runway_months),
        ]

        ws_sum.append(["Metrik Keuangan", "Nilai / Saldo"])
        header_row_idx = 4
        for col in range(1, 3):
            cell = ws_sum.cell(row=header_row_idx, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left" if col == 1 else "right")

        for label, val in summary_rows:
            ws_sum.append([label, val])
            curr_row = ws_sum.max_row
            ws_sum.cell(row=curr_row, column=1).font = bold_font
            ws_sum.cell(row=curr_row, column=1).border = border_thin
            val_cell = ws_sum.cell(row=curr_row, column=2)
            val_cell.font = regular_font
            val_cell.border = border_thin
            if isinstance(val, (int, float)) and "Runway" not in label:
                val_cell.number_format = "#,##0"

        # --- SHEET 2: Histori Transaksi ---
        ws_tx = wb.create_sheet(title="Histori Transaksi")
        ws_tx.views.sheetView[0].showGridLines = True

        tx_headers = ["ID", "Waktu Transaksi", "Tipe", "Kategori", "Nominal", "Merchant / Klien", "Catatan"]
        ws_tx.append(tx_headers)
        for col in range(1, len(tx_headers) + 1):
            cell = ws_tx.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for tx in transactions:
            row_data = [
                tx.id,
                tx.timestamp.strftime("%Y-%m-%d %H:%M"),
                tx.type.value,
                tx.category.value,
                tx.amount,
                tx.source_or_merchant,
                tx.notes or "",
            ]
            ws_tx.append(row_data)
            curr_row = ws_tx.max_row
            for col_idx in range(1, len(row_data) + 1):
                cell = ws_tx.cell(row=curr_row, column=col_idx)
                cell.font = regular_font
                cell.border = border_thin
                if col_idx == 5:
                    cell.number_format = "#,##0"

        # --- SHEET 3: Daftar Invoice & Piutang ---
        ws_inv = wb.create_sheet(title="Daftar Invoice")
        ws_inv.views.sheetView[0].showGridLines = True

        inv_headers = [
            "No. Invoice",
            "Klien",
            "Email Klien",
            "Nama Proyek",
            "Total Tagihan",
            "Mata Uang",
            "Tgl Terbit",
            "Jatuh Tempo",
            "Status",
            "Tgl Lunas",
            "Catatan",
        ]
        ws_inv.append(inv_headers)
        for col in range(1, len(inv_headers) + 1):
            cell = ws_inv.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for inv in invoices:
            row_data = [
                inv.id,
                inv.client_name,
                inv.client_email or "",
                inv.project_title,
                inv.amount,
                inv.currency,
                inv.issue_date,
                inv.due_date,
                inv.status.value,
                inv.paid_date or "-",
                inv.notes or "",
            ]
            ws_inv.append(row_data)
            curr_row = ws_inv.max_row
            for col_idx in range(1, len(row_data) + 1):
                cell = ws_inv.cell(row=curr_row, column=col_idx)
                cell.font = regular_font
                cell.border = border_thin
                if col_idx == 5:
                    cell.number_format = "#,##0"

        # Auto-adjust column widths for all sheets
        for sheet in [ws_sum, ws_tx, ws_inv]:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or "")
                    max_len = max(max_len, len(val_str))
                sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Remove initial empty default sheet
        if default_sheet in wb.worksheets:
            wb.remove(default_sheet)

        # Save to buffer
        out_buf = io.BytesIO()
        wb.save(out_buf)
        out_bytes = out_buf.getvalue()
        out_buf.close()
        return out_bytes


excel_exporter = ExcelExporter()
